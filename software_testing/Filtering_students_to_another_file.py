import os
from openpyxl import load_workbook,Workbook
if os.path.exists("studentdetails.xlsx"):
    wb1= load_workbook("studentdetails.xlsx")
    ws1= wb1.worksheets[0]
else:
    print("File doesn't exist")
    exit()

wb2 = Workbook()
ws2 = wb2.active
rows=ws1.max_row
cols=ws1.max_column
for r in range(2,rows+1):
    cnt=0
    lst=[]
    lst.append(ws1.cell(row=r,column=1).value)
    lst.append(ws1.cell(row=r,column=2).value)
    for c in range(3,cols+1):
        lst.append(ws1.cell(row=r,column=c).value)
        if(int(ws1.cell(row=r,column=c).value)>=20):
            cnt+=1
    if(cnt<3):
        ws2.append(lst)

wb1.save("studentdetails.xlsx")
wb2.save("Fail_list.xlsx")
os.system("studentdetails.xlsx")
os.system("Fail_list.xlsx")
